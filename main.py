import os
import json
import re
import asyncio
import logging
import tempfile
import base64
from datetime import datetime
from typing import Optional

import httpx
import PyPDF2
from aiogram import Bot, Dispatcher, Router, F, types
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    BotCommand,
    BotCommandScopeAllPrivateChats,
    BotCommandScopeAllGroupChats,
    BotCommandScopeChat,
)
from aiogram.enums import ChatType, ParseMode

# ─── Configuration ───────────────────────────────────────────────────────────

BOT_TOKEN = "5278733059:AAGfcKBZF-tA92iXykMIUMKpBzvln6LuQjs"
OWNER_ID = 5360075159

PIONEER_API_URL = "https://api.pioneer.ai/v1/chat/completions"

# Set your proxy URL here if your VPS IP is blocked
# Example: "http://user:pass@proxy:port" or "socks5://user:pass@proxy:port"
# Leave empty or None to disable proxy
PROXY_URL = ""

MODELS = {
    "opus": {
        "name": "Claude Opus 4",
        "model": "claude-opus-4-7",
        "api_key": "pio_sk_2000c121-00ea-41a7-b6b5-4bbb8bdae1de_d0lph1n_WYb3YhCgo-BfBZX2",
    },
    "gpt": {
        "name": "GPT 5.5",
        "model": "gpt-5.5",
        "api_key": "pio_sk_2000c121-00ea-41a7-b6b5-4bbb8bdae1de_d0lph1n_L9QvrTiqT1glmhq4",
    },
    "qwen": {
        "name": "Qwen3 8B",
        "model": "Qwen/Qwen3-8B",
        "api_key": "pio_sk_2000c121-00ea-41a7-b6b5-4bbb8bdae1de_w0lf_8QF1tSC39OtWdKbH",
    },
    "gemini": {
        "name": "Gemini 3.5 Flash",
        "model": "gemini-3.5-flash",
        "api_key": "pio_sk_2000c121-00ea-41a7-b6b5-4bbb8bdae1de_t1g3r_gUVH1wHDkQIfeqae",
    },
}

DEFAULT_MODEL = "gpt"
MAX_HISTORY = 20
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

SYSTEM_PROMPT = (
    "You are a helpful AI assistant\\. Format your replies using Telegram MarkdownV2\\.\n\n"
    "Use these formatting options:\n"
    "*bold* for headings and key points \\(single asterisk, NOT double\\)\n"
    "_italic_ for emphasis\n"
    "`inline code` for code and commands\n"
    "```\\ncode block\\n``` for multi\\-line code\n"
    "> blockquote for quotes\n\n"
    "IMPORTANT: In normal text, these characters must be escaped with \\\\\n"
    "Characters: \\_ \\* \\[ \\] \\( \\) \\~ \\` \\> \\# \\+ \\- \\= \\| \\{ \\} \\. \\!\n\n"
    "Do NOT use:\n"
    "\\- Double asterisks \\*\\*bold\\*\\* \\(use single \\*bold\\*\\)\n"
    "\\- HTML tags\n"
    "\\- Emoji\n"
    "Keep replies concise and well\\-structured\\.\n\n"
    "FILE CREATION:\n"
    "When the user asks you to create a file, generate the content and wrap it like this:\n"
    "[FILE: filename\\.ext]\n"
    "file content here\n"
    "[/FILE]\n\n"
    "Supported file types:\\.txt \\.py \\.js \\.html \\.css \\.json \\.xml \\.yaml \\.md \\.csv \\.sql \\.java \\.c \\.cpp \\.go \\.rs \\.rb \\.php \\.sh \\.pdf \\.docx \\.xlsx and more\n\n"
    "For PDF files, use \\.pdf extension and provide plain text content \\(the bot will convert it to PDF\\)\n"
    "You can create multiple files\\. Always include the full file content\\.\n"
    "After the file block, explain what the file does\\."
)

TELEGRAPH_API = "https://api.graph.org"

# MarkdownV2 escape helpers
_ESC = "\\"
_DOT = "\\."
_EXC = "\\!"
_DASH = "\\-"
_GT = "\\>"
_LPAREN = "\\("
_RPAREN = "\\)"
_HASH = "\\#"
_PLUS = "\\+"

# ─── Logging ─────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("yasir-bot")

# ─── Data Persistence ────────────────────────────────────────────────────────


def get_default_data() -> dict:
    return {
        "owner_id": OWNER_ID,
        "authorized_groups": {},
        "group_models": {},
        "conversations": {},
        "telegraph_token": None,
    }


def load_data() -> dict:
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                defaults = get_default_data()
                for key in defaults:
                    data.setdefault(key, defaults[key])
                return data
        except (json.JSONDecodeError, IOError):
            logger.warning("data.json corrupted or unreadable, creating fresh copy.")
    return get_default_data()


def save_data(data: dict):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


db = load_data()


# ─── Telegraph Integration ───────────────────────────────────────────────────


async def get_telegraph_token() -> str:
    if db.get("telegraph_token"):
        return db["telegraph_token"]
    client_kwargs = {}
    if PROXY_URL:
        client_kwargs["proxy"] = PROXY_URL
    async with httpx.AsyncClient(**client_kwargs) as client:
        resp = await client.post(
            f"{TELEGRAPH_API}/createAccount",
            data={"short_name": "AI_Bot", "author_name": "AI Bot"},
        )
        data = resp.json()
        token = data["result"]["access_token"]
        db["telegraph_token"] = token
        save_data(db)
        return token


async def create_telegraph_page(title: str, content: str, author: str = "AI Bot") -> str:
    token = await get_telegraph_token()

    content_nodes = []
    for line in content.split("\n"):
        if line.strip():
            content_nodes.append({"tag": "p", "children": [line]})

    if not content_nodes:
        content_nodes.append({"tag": "p", "children": [content]})

    client_kwargs = {}
    if PROXY_URL:
        client_kwargs["proxy"] = PROXY_URL
    async with httpx.AsyncClient(**client_kwargs) as client:
        resp = await client.post(
            f"{TELEGRAPH_API}/createPage",
            data={
                "access_token": token,
                "title": title[:256],
                "content": json.dumps(content_nodes),
                "author_name": author,
            },
        )
        data = resp.json()
        return data["result"]["url"]


def extract_code_blocks(text: str) -> list[tuple[str, str]]:
    """Extract ```lang ... ``` blocks from text. Returns [(lang, code), ...]."""
    blocks = []
    pattern = r"```(\w*)\n(.*?)```"
    for match in re.finditer(pattern, text, re.DOTALL):
        lang = match.group(1) or "text"
        code = match.group(2).strip()
        blocks.append((lang, code))
    return blocks


# ─── AI API Call ─────────────────────────────────────────────────────────────


async def call_ai(messages: list[dict], model_key: str) -> str:
    model_info = MODELS.get(model_key, MODELS[DEFAULT_MODEL])
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {model_info['api_key']}",
        "User-Agent": "TelegramBot/1.0",
    }
    payload = {
        "model": model_info["model"],
        "messages": messages,
        "max_tokens": 4096,
        "temperature": 0.7,
        "stream": False,
    }

    client_kwargs = {"timeout": 180}
    if PROXY_URL:
        client_kwargs["proxy"] = PROXY_URL

    for attempt in range(3):
        try:
            async with httpx.AsyncClient(**client_kwargs) as client:
                resp = await client.post(PIONEER_API_URL, json=payload, headers=headers)
                if resp.status_code == 200:
                    data = resp.json()
                    return data["choices"][0]["message"]["content"].strip()
                elif resp.status_code in (403, 429, 500, 502, 503):
                    logger.warning(f"AI API error {resp.status_code}, attempt {attempt + 1}/3")
                    if attempt < 2:
                        await asyncio.sleep(2 * (attempt + 1))
                        continue
                    return f"AI API error ({resp.status_code}). Please try again later."
                else:
                    logger.error(f"AI API error {resp.status_code}: {resp.text}")
                    return f"AI API error ({resp.status_code}). Please try again later."
        except httpx.TimeoutException:
            if attempt < 2:
                logger.warning(f"AI request timeout, attempt {attempt + 1}/3")
                await asyncio.sleep(2)
                continue
            return "AI request timed out. Please try again."
        except Exception as e:
            logger.error(f"AI call failed: {e}")
            return f"Error calling AI: {e}"

    return "AI API error. Please try again later."


async def download_photo_as_base64(message: Message) -> Optional[str]:
    """Download the largest photo from a message and return as base64."""
    if not message.photo:
        return None
    photo = message.photo[-1]
    file = await bot.get_file(photo.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp_path = tmp.name
        await bot.download_file(file.file_path, tmp_path)
    try:
        with open(tmp_path, "rb") as f:
            image_data = f.read()
        return base64.b64encode(image_data).decode("utf-8")
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# ─── File Reading ────────────────────────────────────────────────────────────

TEXT_EXTENSIONS = {
    ".txt", ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".htm", ".css",
    ".scss", ".sass", ".less", ".json", ".xml", ".yaml", ".yml", ".toml",
    ".ini", ".cfg", ".conf", ".env", ".sh", ".bash", ".zsh", ".bat", ".cmd",
    ".ps1", ".psm1", ".md", ".markdown", ".rst", ".tex", ".log", ".csv",
    ".tsv", ".sql", ".r", ".rb", ".go", ".rs", ".java", ".kt", ".swift",
    ".c", ".cpp", ".h", ".hpp", ".cs", ".fs", ".lua", ".perl", ".pl",
    ".php", ".dart", ".scala", ".groovy", ".ex", ".exs", ".erl", ".hs",
        ".ml", ".clj", ".lisp", ".el", ".vim", ".dockerfile", ".makefile",
    ".cmake", ".gradle", ".properties", ".gitignore", ".dockerignore",
    ".htaccess", ".nginx", ".apache", ".vue", ".svelte", ".astro",
}

BINARY_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
}

ARCHIVE_EXTENSIONS = {
    ".zip", ".tar", ".gz", ".tgz", ".bz2", ".xz", ".7z",
}


async def read_text_file(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, "r", encoding="latin-1") as f:
                return f.read()
        except Exception:
            return "[Could not decode file]"


async def read_pdf_file(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        return f"[Error reading PDF: {e}]"
    return text.strip()


async def read_docx_file(file_path: str) -> str:
    try:
        import docx
        doc = docx.Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as e:
        return f"[Error reading DOCX: {e}]"


async def read_xlsx_file(file_path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        output = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    output.append(row_text)
        wb.close()
        return "\n".join(output)
    except Exception as e:
        return f"[Error reading XLSX: {e}]"


async def read_archive_file(file_path: str) -> str:
    import zipfile
    import tarfile

    output = []
    try:
        if zipfile.is_zipfile(file_path):
            with zipfile.ZipFile(file_path, "r") as zf:
                file_list = zf.namelist()
                output.append(f"ZIP archive with {len(file_list)} files:\n")
                for name in file_list[:50]:
                    output.append(f"  {name}")
                text_files = [
                    n for n in file_list
                    if os.path.splitext(n)[1].lower() in TEXT_EXTENSIONS
                    and not n.startswith("__MACOSX")
                ]
                if text_files:
                    output.append(f"\n--- Content of text files ---\n")
                    for name in text_files[:20]:
                        try:
                            with zf.open(name) as f:
                                content = f.read().decode("utf-8", errors="replace")
                            if len(content) > 2000:
                                content = content[:2000] + "\n[... truncated]"
                            output.append(f"\n>>> {name} <<<\n{content}")
                        except Exception:
                            output.append(f"\n>>> {name} <<< [unreadable]")
        elif tarfile.is_tarfile(file_path):
            with tarfile.open(file_path, "r:*") as tf:
                members = tf.getmembers()
                file_list = [m.name for m in members if m.isfile()]
                output.append(f"TAR archive with {len(file_list)} files:\n")
                for name in file_list[:50]:
                    output.append(f"  {name}")
                text_files = [
                    m for m in members
                    if m.isfile() and os.path.splitext(m.name)[1].lower() in TEXT_EXTENSIONS
                ]
                if text_files:
                    output.append(f"\n--- Content of text files ---\n")
                    for member in text_files[:20]:
                        try:
                            f = tf.extractfile(member)
                            if f:
                                content = f.read().decode("utf-8", errors="replace")
                                if len(content) > 2000:
                                    content = content[:2000] + "\n[... truncated]"
                                output.append(f"\n>>> {member.name} <<<\n{content}")
                        except Exception:
                            output.append(f"\n>>> {member.name} <<< [unreadable]")
        else:
            return "[Archive format not supported. Supported: .zip, .tar, .gz, .tgz, .bz2, .xz]"
    except Exception as e:
        return f"[Error reading archive: {e}]"

    return "\n".join(output) if output else "[Archive is empty]"


async def download_and_read_file(message: Message, document=None) -> Optional[str]:
    doc = document or (message.document if message else None)
    if not doc:
        return None

    file_name = doc.file_name or ""
    ext = os.path.splitext(file_name)[1].lower()

    all_extensions = TEXT_EXTENSIONS | BINARY_EXTENSIONS | ARCHIVE_EXTENSIONS
    if ext not in all_extensions and ext != "":
        return None

    file_size = doc.file_size or 0
    if file_size > 20 * 1024 * 1024:
        return "[File too large. Max 20MB supported.]"

    file = await bot.get_file(doc.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp_path = tmp.name
        await bot.download_file(file.file_path, tmp_path)

    try:
        if ext in TEXT_EXTENSIONS:
            content = await read_text_file(tmp_path)
        elif ext == ".pdf":
            content = await read_pdf_file(tmp_path)
        elif ext in (".docx", ".doc"):
            content = await read_docx_file(tmp_path)
        elif ext in (".xlsx", ".xls"):
            content = await read_xlsx_file(tmp_path)
        elif ext in ARCHIVE_EXTENSIONS:
            content = await read_archive_file(tmp_path)
        else:
            content = await read_text_file(tmp_path)
    except Exception as e:
        content = f"[Error reading file: {e}]"
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    if not content:
        return "[File is empty or unreadable]"

    if len(content) > 12000:
        content = content[:12000] + "\n\n[... truncated due to length]"

    return content


# ─── Conversation History ────────────────────────────────────────────────────


def get_history(chat_id: int) -> list[dict]:
    key = str(chat_id)
    return db["conversations"].get(key, [])


def add_to_history(chat_id: int, role: str, content: str):
    key = str(chat_id)
    if key not in db["conversations"]:
        db["conversations"][key] = []
    db["conversations"][key].append({"role": role, "content": content})
    if len(db["conversations"][key]) > MAX_HISTORY:
        db["conversations"][key] = db["conversations"][key][-MAX_HISTORY:]
    save_data(db)


def clear_history(chat_id: int):
    key = str(chat_id)
    db["conversations"][key] = []
    save_data(db)


# ─── Bot Setup ───────────────────────────────────────────────────────────────

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()
router = Router()
dp.include_router(router)

# ─── Helpers ─────────────────────────────────────────────────────────────────


def is_owner(user_id: int) -> bool:
    return user_id == db.get("owner_id", OWNER_ID)


def is_authorized(chat_id: int) -> bool:
    return str(chat_id) in db["authorized_groups"]


def get_group_model(chat_id: int) -> str:
    return db["group_models"].get(str(chat_id), DEFAULT_MODEL)


def model_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    for key, info in MODELS.items():
        buttons.append(
            [InlineKeyboardButton(text=info["name"], callback_data=f"setmodel:{key}")]
        )
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def is_bot_mentioned(message: Message) -> bool:
    if not message.entities or not message.text:
        return False
    for entity in message.entities:
        if entity.type == "mention":
            mention_text = message.text[entity.offset : entity.offset + entity.length]
            if bot.username and mention_text.lower() == f"@{bot.username.lower()}":
                return True
    return False


def strip_bot_mention(text: str) -> str:
    if bot.username:
        return re.sub(rf"@{re.escape(bot.username)}\s*", "", text, flags=re.IGNORECASE).strip()
    return text


def mentions_other_user(message: Message) -> bool:
    """Check if message mentions any user OTHER than the bot."""
    if not message.entities:
        return False
    for entity in message.entities:
        if entity.type == "text_mention" and entity.user:
            if entity.user.id != bot.id:
                return True
        elif entity.type == "mention":
            mention_text = message.text[entity.offset : entity.offset + entity.length]
            if bot.username and mention_text.lower() != f"@{bot.username.lower()}":
                return True
    return False


def md_escape(text: str) -> str:
    special = r"_*[]()~`>#+-=|{}.!\\"
    result = []
    for ch in text:
        if ch in special:
            result.append("\\")
        result.append(ch)
    return "".join(result)


def md_to_telegram(text: str) -> str:
    """Convert standard Markdown to Telegram MarkdownV2 and escape special chars."""

    text = re.sub(r'</?b>', '*', text)
    text = re.sub(r'</?strong>', '*', text)
    text = re.sub(r'</?i>', '_', text)
    text = re.sub(r'</?em>', '_', text)
    text = re.sub(r'</?code>', '`', text)
    text = re.sub(r'</?pre>', '```', text)
    text = re.sub(r'</?blockquote>', '', text)
    text = re.sub(r'<br\s*/?>', '\n', text)
    text = re.sub(r'</?\w+[^>]*>', '', text)

    def convert_bold(m):
        inner = m.group(1)
        return '*' + inner + '*'
    text = re.sub(r'\*\*(.+?)\*\*', convert_bold, text)

    text = re.sub(r'^(\s*)-\s', r'\1\- ', text, flags=re.MULTILINE)
    text = re.sub(r'^(\s*)\*\s', r'\1\- ', text, flags=re.MULTILINE)

    def convert_blockquote(m):
        inner = m.group(1)
        return '>' + inner
    text = re.sub(r'^>\s?(.*)$', convert_blockquote, text, flags=re.MULTILINE)

    text = re.sub(r'~~(.+?)~~', r'~\1~', text)

    text = text.replace(' :)', ' \\:\\)')
    text = text.replace(' :(', ' \\:\\(')

    special_chars = set(r"_*[]()~`>#+-=|{}.!")
    parts = []
    i = 0
    in_code_block = False
    in_inline_code = False

    while i < len(text):
        if not in_code_block and not in_inline_code and text[i:i+3] == "```":
            parts.append("```")
            in_code_block = True
            i += 3
            continue

        if in_code_block and text[i:i+3] == "```":
            parts.append("```")
            in_code_block = False
            i += 3
            continue

        if not in_code_block and not in_inline_code and text[i] == "`":
            parts.append("`")
            in_inline_code = True
            i += 1
            continue

        if in_inline_code and text[i] == "`":
            parts.append("`")
            in_inline_code = False
            i += 1
            continue

        if in_code_block or in_inline_code:
            parts.append(text[i])
            i += 1
            continue

        if text[i] == "\\" and i + 1 < len(text):
            parts.append(text[i:i+2])
            i += 2
            continue

        if text[i] in special_chars:
            parts.append("\\")
            parts.append(text[i])
            i += 1
            continue

        parts.append(text[i])
        i += 1

    return "".join(parts)


def get_user_mention(user) -> str:
    name = user.first_name or user.username or "User"
    escaped_name = md_escape(name)
    return f"[{escaped_name}](tg://user?id={user.id})"


def split_text(text: str, max_len: int = 4000) -> list[str]:
    if len(text) <= max_len:
        return [text]

    parts = []
    current = ""

    paragraphs = text.split("\n\n")
    for para in paragraphs:
        if len(current) + len(para) + 2 <= max_len:
            current = f"{current}\n\n{para}" if current else para
        else:
            if current:
                parts.append(current)
            if len(para) <= max_len:
                current = para
            else:
                lines = para.split("\n")
                current = ""
                for line in lines:
                    if len(current) + len(line) + 1 <= max_len:
                        current = f"{current}\n{line}" if current else line
                    else:
                        if current:
                            parts.append(current)
                        if len(line) <= max_len:
                            current = line
                        else:
                            while len(line) > max_len:
                                parts.append(line[:max_len])
                                line = line[max_len:]
                            current = line

    if current:
        parts.append(current)

    return parts


def extract_files(text: str) -> tuple[list[tuple[str, str]], str]:
    """Extract [FILE: name]...[/FILE] blocks from AI response.
    Returns (list of (filename, content), remaining text)."""
    files = []
    pattern = r'\[FILE:\s*(.+?)\]\s*\n(.*?)\n\[/FILE\]'
    matches = list(re.finditer(pattern, text, re.DOTALL))

    for match in matches:
        filename = match.group(1).strip()
        content = match.group(2)
        content = strip_code_block(content)
        files.append((filename, content))

    remaining = re.sub(pattern, '', text, flags=re.DOTALL).strip()
    return files, remaining


def strip_code_block(text: str) -> str:
    """Remove ```lang ... ``` wrapping from code."""
    text = text.strip()
    if text.startswith("```") and text.endswith("```"):
        first_line_end = text.index("\n") if "\n" in text else len(text)
        text = text[first_line_end + 1:]
        text = text[:-3]
    return text.strip()


def create_pdf(content: str, title: str = "Document") -> bytes:
    """Create a PDF from text content."""
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    try:
        pdf.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", uni=True)
        pdf.set_font("DejaVu", size=11)
    except Exception:
        try:
            pdf.add_font("Arial", "", "C:\\Windows\\Fonts\\arial.ttf", uni=True)
            pdf.set_font("Arial", size=11)
        except Exception:
            pdf.set_font("Helvetica", size=11)

    pdf.set_font_size(16)
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)
    pdf.set_font_size(11)

    for line in content.split("\n"):
        try:
            pdf.multi_cell(0, 6, line)
        except Exception:
            safe_line = line.encode("ascii", "replace").decode()
            pdf.multi_cell(0, 6, safe_line)

    return bytes(pdf.output())


async def send_files(message: Message, files: list[tuple[str, str]]):
    """Create and send files to the user."""
    for filename, content in files:
        safe_name = re.sub(r'[^\w\-.]', '_', filename)
        ext = os.path.splitext(safe_name)[1].lower()
        tmp_path = os.path.join(tempfile.gettempdir(), f"bot_{safe_name}")

        try:
            if ext == ".pdf":
                title = os.path.splitext(safe_name)[0]
                pdf_bytes = create_pdf(content, title)
                with open(tmp_path, 'wb') as f:
                    f.write(pdf_bytes)
            else:
                with open(tmp_path, 'w', encoding='utf-8') as f:
                    f.write(content)

            await message.reply_document(
                types.FSInputFile(tmp_path, filename=safe_name),
                caption=f"Created `{md_escape(safe_name)}`",
                parse_mode=ParseMode.MARKDOWN_V2,
            )
        except Exception as e:
            await message.reply(f"Failed to create `{md_escape(safe_name)}`: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


async def send_ai_response(message: Message, user_text: str, model_key: str, chat_id: int):
    thinking_msg = await message.reply(f"Thinking{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)

    history = get_history(chat_id)
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    messages.extend(history)
    messages.append({"role": "user", "content": user_text})

    add_to_history(chat_id, "user", user_text)

    response = await call_ai(messages, model_key)
    add_to_history(chat_id, "assistant", response)

    try:
        await thinking_msg.delete()
    except Exception:
        pass

    files, remaining = extract_files(response)

    if files:
        await send_files(message, files)

    if not remaining and files:
        return

    if not remaining:
        remaining = response

    response = md_to_telegram(remaining)

    mention = ""
    if message.from_user:
        mention = get_user_mention(message.from_user)

    parts = split_text(response, max_len=4000)

    for i, part in enumerate(parts):
        if i == 0 and mention:
            text = f"{mention}\n\n{part}"
        else:
            text = part

        try:
            if i == 0:
                await message.reply(text, parse_mode=ParseMode.MARKDOWN_V2)
            else:
                await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)
        except Exception:
            if i == 0:
                await message.reply(part)
            else:
                await message.answer(part)


# ─── Command Handlers ────────────────────────────────────────────────────────


@router.message(CommandStart())
async def cmd_start(message: Message):
    text = (
        f"Welcome{_EXC} I'm your *AI Group Manager Bot*{_DOT}\n\n"
        f"Add me to a group and use /authorize to activate me there{_DOT}\n"
        f"Use /settings to change the AI model{_DOT}\n"
        f"Use /q to ask me anything{_DOT}\n\n"
        f"> contact:@itznik\\_x"
    )
    await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("authorize"))
async def cmd_authorize(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can authorize groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    chat_id = str(message.chat.id)
    chat_title = md_escape(message.chat.title or chat_id)

    if chat_id in db["authorized_groups"]:
        await message.answer(f"Group *{chat_title}* is already authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    db["authorized_groups"][chat_id] = {
        "title": message.chat.title or chat_id,
        "authorized_at": datetime.now().isoformat(),
        "authorized_by": message.from_user.id,
    }
    db["group_models"][chat_id] = DEFAULT_MODEL
    save_data(db)

    model_name = md_escape(MODELS[DEFAULT_MODEL]["name"])
    await message.answer(
        f"Group *{chat_title}* has been authorized{_EXC}\n"
        f"Current AI model: *{model_name}*\n"
        f"Use /settings to change the model{_DOT}",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.message(Command("deauthorize", "unauthorize"))
async def cmd_deauthorize(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can deauthorize groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    chat_id = str(message.chat.id)
    chat_title = md_escape(message.chat.title or chat_id)

    if chat_id not in db["authorized_groups"]:
        await message.answer(f"Group *{chat_title}* is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    del db["authorized_groups"][chat_id]
    db["group_models"].pop(chat_id, None)
    db["conversations"].pop(chat_id, None)
    save_data(db)

    await message.answer(f"Group *{chat_title}* has been deauthorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("settings"))
async def cmd_settings(message: Message):
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can change settings{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if message.chat.type == ChatType.PRIVATE:
        await message.answer(
            f"Settings can only be changed inside a group{_DOT}\n"
            f"Go to your group and use /settings there{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return

    chat_id = str(message.chat.id)
    if chat_id not in db["authorized_groups"]:
        await message.answer(f"This group is not authorized{_DOT} Use /authorize first{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    current_model = get_group_model(message.chat.id)
    current_name = md_escape(MODELS[current_model]["name"])

    await message.answer(
        f"Current model: *{current_name}*\n\nSelect a new AI model:",
        reply_markup=model_keyboard(),
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.callback_query(F.data.startswith("setmodel:"))
async def cb_set_model(callback: CallbackQuery):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can change models.", show_alert=True)
        return

    if callback.message.chat.type == ChatType.PRIVATE:
        await callback.answer("Change model inside a group.", show_alert=True)
        return

    model_key = callback.data.split(":")[1]
    if model_key not in MODELS:
        await callback.answer("Invalid model.", show_alert=True)
        return

    chat_id = str(callback.message.chat.id)
    if chat_id not in db["authorized_groups"]:
        await callback.answer("Group not authorized.", show_alert=True)
        return

    db["group_models"][chat_id] = model_key
    save_data(db)

    model_name = md_escape(MODELS[model_key]["name"])
    await callback.message.edit_text(
        f"AI model changed to *{model_name}*",
        parse_mode=ParseMode.MARKDOWN_V2,
    )
    await callback.answer(f"Model set to {MODELS[model_key]['name']}")


@router.message(Command("clearhistory"))
async def cmd_clearhistory(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can clear history{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    clear_history(message.chat.id)
    await message.answer(f"Conversation history cleared for this group{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("model"))
async def cmd_model(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_authorized(message.chat.id):
        await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    current = get_group_model(message.chat.id)
    name = md_escape(MODELS[current]["name"])
    await message.answer(f"Current AI model: *{name}*", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("q", "question"))
async def cmd_question(message: Message):
    if not message.text:
        return
    question_text = message.text.split(None, 1)
    if len(question_text) < 2 or not question_text[1].strip():
        await message.answer("Usage: `/q your question here`", parse_mode=ParseMode.MARKDOWN_V2)
        return

    user_text = question_text[1].strip()

    if message.reply_to_message:
        replied = message.reply_to_message

        if replied.document:
            file_content = await download_and_read_file(message, document=replied.document)
            if file_content:
                file_name = replied.document.file_name or "file"
                user_text = f"User asked: {user_text}\n\n--- File: {file_name} ---\n{file_content}"

        elif replied.photo:
            user_text = f"User asked about an image: {user_text}"

        elif replied.text:
            replied_text = replied.text
            user_text = f"User asked: {user_text}\n\n--- Replied message ---\n{replied_text}"

        elif replied.caption:
            replied_text = replied.caption
            user_text = f"User asked: {user_text}\n\n--- Replied message ---\n{replied_text}"

    if message.chat.type == ChatType.PRIVATE:
        await send_ai_response(message, user_text, DEFAULT_MODEL, message.chat.id)
        return

    if not is_authorized(message.chat.id):
        await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.chat.id)


@router.message(Command("post"))
async def cmd_post(message: Message):
    content = None
    title = "Shared Content"

    if message.reply_to_message:
        replied = message.reply_to_message
        if replied.document:
            content = await download_and_read_file(message, document=replied.document)
            title = replied.document.file_name or "File Content"
        elif replied.text:
            content = replied.text
            title = replied.text[:50].split("\n")[0] or "Shared Text"
        elif replied.caption:
            content = replied.caption
            title = "Shared Content"

    elif message.text:
        parts = message.text.split(None, 1)
        if len(parts) >= 2 and parts[1].strip():
            content = parts[1].strip()
            title = content[:50].split("\n")[0] or "Shared Text"

    if not content:
        await message.answer(
            f"Usage: reply to a message with /post or /post your text here",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return

    try:
        thinking = await message.reply(f"Publishing to Telegraph{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        url = await create_telegraph_page(title, content)
        try:
            await thinking.delete()
        except Exception:
            pass
        author = message.from_user.first_name if message.from_user else "User"
        escaped_author = md_escape(author)
        escaped_url = md_escape(url)
        await message.reply(
            f"Published to Telegraph by *{escaped_author}*{_DOT}\n\n{escaped_url}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
    except Exception as e:
        await message.reply(f"Failed to publish: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("help"))
async def cmd_help(message: Message):
    text = (
        f"*AI Group Manager Bot {_DASH} Commands*\n\n"
        f"/start {_DASH} Welcome message\n"
        f"/q `question` {_DASH} Ask AI a question directly\n"
        f"/post {_DASH} Publish text or code to Telegraph\n"
        f"/authorize {_DASH} Activate bot in this group {_LPAREN}owner only{_RPAREN}\n"
        f"/deauthorize or /unauthorize {_DASH} Deactivate bot {_LPAREN}owner only{_RPAREN}\n"
        f"/settings {_DASH} Change AI model {_LPAREN}owner only{_RPAREN}\n"
        f"/model {_DASH} Show current AI model\n"
        f"/clearhistory {_DASH} Clear conversation history {_LPAREN}owner only{_RPAREN}\n"
        f"/help {_DASH} Show this help message\n\n"
        f"*How to use:*\n"
        f"{_DASH} Reply to my message to chat with me\n"
        f"{_DASH} Tag me with @ to ask something\n"
        f"{_DASH} Send any file to read it {_LPAREN}txt, pdf, docx, xlsx, zip, py, js, etc{_RPAREN}\n"
        f"{_DASH} Reply to any message or file with /q\n"
        f"{_DASH} Use /post to publish text or code to Telegraph\n"
        f"{_DASH} Use /q for direct questions"
    )
    await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)


# ─── Document Handler (File Reading) ─────────────────────────────────────────


@router.message(F.document)
async def handle_document(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        file_content = await download_and_read_file(message)
        if file_content is None:
            await message.answer(f"File format not supported{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
            return

        caption = message.caption or "Summarize and explain the content of this file."
        user_text = f"{caption}\n\n--- File Content ---\n{file_content}"

        await send_ai_response(message, user_text, DEFAULT_MODEL, message.chat.id)
        return

    if not is_authorized(message.chat.id):
        return

    is_reply_to_bot = (
        message.reply_to_message
        and message.reply_to_message.from_user
        and message.reply_to_message.from_user.id == bot.id
    )
    is_mentioned = is_bot_mentioned(message)

    if not is_reply_to_bot and not is_mentioned:
        return

    if is_reply_to_bot and mentions_other_user(message):
        return

    file_content = await download_and_read_file(message)
    if file_content is None:
        await message.answer(f"File format not supported{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    caption = message.caption or "Summarize and explain the content of this file."
    user_text = f"{caption}\n\n--- File Content ---\n{file_content}"

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.chat.id)


@router.message(F.photo)
async def handle_photo(message: Message):
    await message.answer(
        f"Image analysis is not supported with the current API{_DOT}\n"
        f"Use /q with a description of the image instead{_DOT}",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


# ─── Message Handler (AI Reply) ──────────────────────────────────────────────


@router.message(F.text & ~F.text.startswith("/"))
async def handle_message(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        user_text = message.text
        await send_ai_response(message, user_text, DEFAULT_MODEL, message.chat.id)
        return

    if not is_authorized(message.chat.id):
        return

    is_reply_to_bot = (
        message.reply_to_message
        and message.reply_to_message.from_user
        and message.reply_to_message.from_user.id == bot.id
    )
    is_mentioned = is_bot_mentioned(message)

    if not is_reply_to_bot and not is_mentioned:
        return

    if is_reply_to_bot and mentions_other_user(message):
        return

    user_text = strip_bot_mention(message.text) if is_mentioned else message.text
    if not user_text:
        user_text = "Hello"

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.chat.id)


# ─── Startup ─────────────────────────────────────────────────────────────────


async def on_startup():
    await bot.delete_webhook(drop_pending_updates=True)
    me = await bot.get_me()
    bot.username = me.username
    logger.info(f"Bot started: @{me.username} (id={me.id})")

    user_commands = [
        BotCommand(command="start", description="Start the bot"),
        BotCommand(command="q", description="Ask AI a question"),
        BotCommand(command="post", description="Publish to Telegraph"),
        BotCommand(command="help", description="Show help"),
        BotCommand(command="model", description="Show current AI model"),
    ]
    owner_commands = [
        BotCommand(command="start", description="Start the bot"),
        BotCommand(command="q", description="Ask AI a question"),
        BotCommand(command="post", description="Publish to Telegraph"),
        BotCommand(command="authorize", description="Activate bot in this group"),
        BotCommand(command="deauthorize", description="Deactivate bot in this group"),
        BotCommand(command="unauthorize", description="Deactivate bot in this group"),
        BotCommand(command="settings", description="Change AI model"),
        BotCommand(command="model", description="Show current AI model"),
        BotCommand(command="clearhistory", description="Clear conversation history"),
        BotCommand(command="help", description="Show help"),
    ]

    await bot.set_my_commands(user_commands, scope=BotCommandScopeAllPrivateChats())
    await bot.set_my_commands(user_commands, scope=BotCommandScopeAllGroupChats())
    await bot.set_my_commands(owner_commands, scope=BotCommandScopeChat(chat_id=OWNER_ID))

    logger.info(f"Owner ID: {db.get('owner_id', OWNER_ID)}")
    logger.info(f"Authorized groups: {len(db['authorized_groups'])}")
    logger.info("Bot is running!")


async def main():
    dp.startup.register(on_startup)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
